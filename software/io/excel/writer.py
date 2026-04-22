"""Excel 文件写入模块。

导出样本数据到 Excel 文件。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING, List

if TYPE_CHECKING:
    from openpyxl import Workbook as WorkbookType
    from openpyxl.styles import Font as FontType, PatternFill as PatternFillType

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None  # type: ignore
    Workbook = None  # type: ignore

from software.io.excel.schema import SampleRow


class ExcelWriter:
    """Excel 文件写入器。"""
    
    def write_samples(
        self,
        samples: List[SampleRow],
        output_path: str,
        *,
        include_status: bool = False,
        include_error: bool = False,
    ) -> None:
        """将样本写入 Excel 文件。
        
        Args:
            samples: 样本列表
            output_path: 输出文件路径
            include_status: 是否包含状态列
            include_error: 是否包含错误信息列
        """
        if openpyxl is None or Workbook is None:
            raise ImportError("需要安装 openpyxl 库：pip install openpyxl")
        
        if not samples:
            logging.warning("没有样本数据，跳过写入")
            return
        
        # 创建工作簿
        wb: WorkbookType = Workbook()  # type: ignore
        ws = wb.active
        if ws is None:
            raise RuntimeError("无法创建工作表")
        ws.title = "未完成样本"
        
        # 获取列名（从第一个样本）
        first_sample = samples[0]
        col_names = list(first_sample.values.keys())
        
        # 添加额外列
        if include_status:
            col_names.append("状态")
        if include_error:
            col_names.append("错误信息")
        
        # 导入样式类
        from openpyxl.styles import Font, PatternFill
        
        # 写入表头
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        
        # 写入数据
        for row_idx, sample in enumerate(samples, start=2):
            # 写入原始数据
            for col_idx, col_name in enumerate(list(sample.values.keys()), start=1):
                value = sample.values.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 写入状态
            if include_status:
                status_col = len(sample.values) + 1
                status_text = self._get_status_text(sample.status)
                cell = ws.cell(row=row_idx, column=status_col, value=status_text)
                # 根据状态设置颜色
                if sample.status == "failed":
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif sample.status == "pending":
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # 写入错误信息
            if include_error:
                error_col = len(sample.values) + (1 if include_status else 0) + 1
                error_text = sample.error or ""
                ws.cell(row=row_idx, column=error_col, value=error_text)
        
        # 自动调整列宽
        from openpyxl.utils import get_column_letter
        
        for col_idx in range(1, len(col_names) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, len(samples) + 2):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logging.info(f"已导出 {len(samples)} 条样本到: {output_path}")
    
    def _get_status_text(self, status: str) -> str:
        """获取状态文本。"""
        status_map = {
            "pending": "待处理",
            "assigned": "处理中",
            "success": "成功",
            "failed": "失败",
        }
        return status_map.get(status, status)
    
    def export_failed_samples(
        self,
        samples: List[SampleRow],
        original_path: str,
    ) -> str:
        """导出失败的样本。
        
        Args:
            samples: 所有样本列表
            original_path: 原始 Excel 文件路径
            
        Returns:
            导出文件路径
        """
        # 筛选失败和待处理的样本
        failed_samples = [
            s for s in samples
            if s.status in ("failed", "pending")
        ]
        
        if not failed_samples:
            logging.info("没有失败或未完成的样本，无需导出")
            return ""
        
        # 生成输出文件名
        original_path_obj = Path(original_path)
        output_filename = f"{original_path_obj.stem}_未完成{original_path_obj.suffix}"
        output_path = str(original_path_obj.parent / output_filename)
        
        # 写入文件
        self.write_samples(
            failed_samples,
            output_path,
            include_status=True,
            include_error=True,
        )
        
        return output_path
    
    def export_validation_failed_samples(
        self,
        samples: List[SampleRow],
        original_path: str,
    ) -> str:
        """导出验证失败的样本（解析失败的数据）。
        
        Args:
            samples: 所有样本列表
            original_path: 原始 Excel 文件路径
            
        Returns:
            导出文件路径
        """
        # 筛选验证失败的样本（status 为 failed 且 error 包含"验证"或"解析"）
        validation_failed = [
            s for s in samples
            if s.status == "failed" and s.error and (
                "验证" in s.error or
                "解析" in s.error or
                "标准化" in s.error or
                "映射" in s.error
            )
        ]
        
        if not validation_failed:
            logging.info("没有验证失败的样本，无需导出")
            return ""
        
        # 生成输出文件名
        original_path_obj = Path(original_path)
        output_filename = f"{original_path_obj.stem}_解析失败{original_path_obj.suffix}"
        output_path = str(original_path_obj.parent / output_filename)
        
        # 写入文件
        self.write_samples(
            validation_failed,
            output_path,
            include_status=True,
            include_error=True,
        )
        
        return output_path

