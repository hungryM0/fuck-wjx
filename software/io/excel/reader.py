"""Excel 文件读取器。"""
from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from software.io.excel.schema import SampleRow

from software.io.excel.schema import SampleRow as _SampleRow


class ExcelReader:
    """Excel 文件读取器。
    
    读取 Excel 文件，第一行为列头（题目标题），后续行为样本数据。
    """

    def read(self, file_path: str) -> list[SampleRow]:
        """读取 Excel 文件。
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            样本行列表
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式错误
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel 文件不存在: {file_path}")
        except Exception as e:
            raise ValueError(f"无法打开 Excel 文件: {e}")

        ws = wb.active
        if ws is None:
            raise ValueError("Excel 文件没有活动工作表")

        # 读取列头（第一行）
        headers = []
        for cell in ws[1]:
            value = cell.value
            if value is None or str(value).strip() == "":
                continue
            headers.append(str(value).strip())

        if not headers:
            raise ValueError("Excel 文件第一行没有列头")

        # 读取数据行
        samples = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # 构建列名到值的映射
            values = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    cell_value = row[i]
                    # 保留原始值，包括 None
                    values[header] = cell_value
                else:
                    values[header] = None

            samples.append(_SampleRow(row_no=row_idx, values=values))

        wb.close()

        if not samples:
            raise ValueError("Excel 文件没有数据行")

        return samples
